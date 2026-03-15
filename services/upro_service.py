"""
UProService — Syndicate management, UPro scores, AOL generation.
"""
import logging
import random
import json
from typing import Optional
import streamlit as st
from services.supabase_client import supabase
from services.base_service import BaseService
from services.grading_service import GradingService, score_to_letter
from config import CACHE_TTL

logger = logging.getLogger("sylemax.upro_service")


# ── Helpers ───────────────────────────────────────────────────────

def _distribute_marks(total: float, parts: int, max_per_part: list[float]) -> list[float]:
    """
    Distribute `total` marks across `parts` items where each item i cannot
    exceed max_per_part[i].  Returns a list of floats (1 decimal place).

    Algorithm:
      1. Clamp total to the sum of all maximums.
      2. Assign each item proportionally: item_i gets (max_i / total_max) * total.
      3. Clamp each item to its individual maximum (handles floating-point drift).
      4. If the sum doesn't exactly equal target due to rounding, redistribute
         the remainder across items that still have headroom — never exceeding max.
    """
    if parts == 0:
        return []

    caps = [float(m) for m in max_per_part]
    total_max = sum(caps)

    if total_max <= 0:
        return [0.0] * parts

    # Clamp total to what is achievable
    target = min(float(total), total_max)
    target = max(0.0, target)

    if target == 0.0:
        return [0.0] * parts

    # Proportional assignment
    result = [round((caps[i] / total_max) * target, 1) for i in range(parts)]

    # Hard-clamp each item to its cap (safety net)
    result = [min(result[i], caps[i]) for i in range(parts)]
    result = [max(0.0, v) for v in result]

    # Fix any drift: redistribute surplus/deficit one tenth at a time
    diff = round(target - sum(result), 1)
    step = 0.1 if diff > 0 else -0.1
    iterations = 0
    while abs(diff) >= 0.05 and iterations < 1000:
        for i in range(parts):
            if abs(diff) < 0.05:
                break
            if step > 0 and result[i] < caps[i]:
                add = min(step, caps[i] - result[i], diff)
                result[i] = round(result[i] + add, 1)
                diff = round(diff - add, 1)
            elif step < 0 and result[i] > 0:
                sub = min(-step, result[i], -diff)
                result[i] = round(result[i] - sub, 1)
                diff = round(diff + sub, 1)
        iterations += 1

    # Final clamp to ensure nothing ever exceeds its cap
    result = [min(round(v, 1), caps[i]) for i, v in enumerate(result)]
    result = [max(0.0, v) for v in result]

    return result


class UProService(BaseService):

    # ── Syndicate CRUD ────────────────────────────────────────────

    @staticmethod
    def get_syndicates(course_uuid: str) -> list:
        try:
            r = supabase.table("syndicates")\
                .select("*, profiles!syndicates_lead_student_id_fkey(id, full_name, first_name, last_name, enrollment_number)")\
                .eq("course_id", course_uuid)\
                .order("name")\
                .execute()
            return r.data or []
        except Exception as e:
            logger.exception(f"Failed to fetch syndicates: {course_uuid}")
            return []

    @staticmethod
    def get_syndicate_members(syndicate_id: str) -> list:
        try:
            r = supabase.table("syndicate_members")\
                .select("*, profiles(id, full_name, first_name, last_name, enrollment_number, program)")\
                .eq("syndicate_id", syndicate_id)\
                .execute()
            return r.data or []
        except Exception as e:
            logger.exception(f"Failed to fetch members: {syndicate_id}")
            return []

    @staticmethod
    def get_student_syndicate(course_uuid: str, student_id: str) -> dict | None:
        """Returns the syndicate a student belongs to in this course."""
        try:
            r = supabase.table("syndicate_members")\
                .select("*, syndicates(*)")\
                .eq("course_id", course_uuid)\
                .eq("student_id", student_id)\
                .execute()
            return r.data[0] if r.data else None
        except Exception:
            return None

    @staticmethod
    def create_syndicate(
        course_uuid: str,
        name: str,
        lead_student_id: str | None,
        created_by_role: str = "faculty_ultra",
        status: str = "confirmed",
    ) -> dict | None:
        try:
            r = supabase.table("syndicates").insert({
                "course_id":       course_uuid,
                "name":            name.strip(),
                "lead_student_id": lead_student_id,
                "created_by_role": created_by_role,
                "status":          status,
            }).execute()
            UProService.clear_cache()
            return r.data[0] if r.data else None
        except Exception as e:
            logger.exception(f"Failed to create syndicate: {name}")
            return None

    @staticmethod
    def update_syndicate(syndicate_id: str, data: dict) -> bool:
        try:
            r = supabase.table("syndicates").update(data)\
                .eq("id", syndicate_id).execute()
            UProService.clear_cache()
            return bool(r.data)
        except Exception as e:
            logger.exception(f"Failed to update syndicate: {syndicate_id}")
            return False

    @staticmethod
    def delete_syndicate(syndicate_id: str) -> bool:
        try:
            supabase.table("syndicates").delete().eq("id", syndicate_id).execute()
            UProService.clear_cache()
            return True
        except Exception as e:
            logger.exception(f"Failed to delete syndicate: {syndicate_id}")
            return False

    @staticmethod
    def confirm_syndicate(syndicate_id: str) -> bool:
        from datetime import datetime, timezone
        return UProService.update_syndicate(syndicate_id, {
            "status": "confirmed",
            "confirmed_at": datetime.now(timezone.utc).isoformat(),
        })

    @staticmethod
    def reject_syndicate(syndicate_id: str) -> bool:
        return UProService.update_syndicate(syndicate_id, {"status": "rejected"})

    @staticmethod
    def add_member(syndicate_id: str, course_uuid: str, student_id: str) -> bool:
        try:
            supabase.table("syndicate_members").upsert({
                "syndicate_id": syndicate_id,
                "student_id":   student_id,
                "course_id":    course_uuid,
            }, on_conflict="syndicate_id,student_id").execute()
            UProService.clear_cache()
            return True
        except Exception as e:
            logger.exception(f"Failed to add member {student_id}")
            return False

    @staticmethod
    def remove_member(syndicate_id: str, student_id: str) -> bool:
        try:
            supabase.table("syndicate_members").delete()\
                .eq("syndicate_id", syndicate_id)\
                .eq("student_id", student_id)\
                .execute()
            UProService.clear_cache()
            return True
        except Exception as e:
            logger.exception(f"Failed to remove member: {student_id}")
            return False

    # ── UPro Scores ───────────────────────────────────────────────

    @staticmethod
    def get_upro_scores(course_uuid: str) -> list:
        try:
            r = supabase.table("upro_scores")\
                .select("*, profiles(id, full_name, first_name, last_name, enrollment_number, program), syndicates(id, name)")\
                .eq("course_id", course_uuid)\
                .execute()
            return r.data or []
        except Exception as e:
            logger.exception(f"Failed to fetch UPro scores: {course_uuid}")
            return []

    @staticmethod
    def get_student_upro_score(course_uuid: str, student_id: str) -> dict | None:
        try:
            r = supabase.table("upro_scores").select("*")\
                .eq("course_id", course_uuid)\
                .eq("student_id", student_id)\
                .execute()
            return r.data[0] if r.data else None
        except Exception:
            return None

    @staticmethod
    def save_upro_score(
        course_uuid: str,
        student_id: str,
        syndicate_id: str | None = None,
        quiz_score: float | None = None,
        assignment_score: float | None = None,
        midterm_score: float | None = None,
        final_score: float | None = None,
        quiz_note: str = "",
        assignment_note: str = "",
        midterm_note: str = "",
        final_note: str = "",
    ) -> bool:
        from datetime import datetime, timezone
        try:
            r = supabase.table("upro_scores").upsert({
                "course_id":        course_uuid,
                "student_id":       student_id,
                "syndicate_id":     syndicate_id,
                "quiz_score":       quiz_score,
                "assignment_score": assignment_score,
                "midterm_score":    midterm_score,
                "final_score":      final_score,
                "quiz_note":        quiz_note,
                "assignment_note":  assignment_note,
                "midterm_note":     midterm_note,
                "final_note":       final_note,
                "updated_at":       datetime.now(timezone.utc).isoformat(),
            }, on_conflict="course_id,student_id").execute()
            return bool(r.data)
        except Exception as e:
            logger.exception(f"Failed to save UPro score: {student_id}")
            return False

    # ── AOL Config ────────────────────────────────────────────────

    @staticmethod
    def get_aol_config(course_uuid: str) -> dict:
        _default = {
            "num_quizzes":           3,
            "quiz_marks_list":       [10.0, 10.0, 10.0],
            "num_assignments":       3,
            "assignment_marks_list": [10.0, 10.0, 10.0],
            "num_midterm_questions": 2,
            "midterm_q_marks":       [12.5, 12.5],
            "num_final_questions":   3,
            "final_q_marks":         [15.0, 15.0, 10.0],
        }
        try:
            r = supabase.table("aol_config").select("*")\
                .eq("course_id", course_uuid).execute()
            if r.data:
                cfg = r.data[0]
                # Parse JSONB arrays
                for key in ["quiz_marks_list", "assignment_marks_list",
                            "midterm_q_marks", "final_q_marks"]:
                    v = cfg.get(key)
                    if isinstance(v, str):
                        try:
                            cfg[key] = json.loads(v)
                        except Exception:
                            cfg[key] = _default[key]
                    elif v is None:
                        cfg[key] = _default[key]
                # Back-compat: if old single-value fields exist, convert to list
                if "quiz_marks_list" not in cfg or not cfg["quiz_marks_list"]:
                    old_max = float(cfg.get("quiz_max_marks") or 10)
                    cfg["quiz_marks_list"] = [old_max] * int(cfg.get("num_quizzes") or 3)
                if "assignment_marks_list" not in cfg or not cfg["assignment_marks_list"]:
                    old_max = float(cfg.get("assignment_max_marks") or 10)
                    cfg["assignment_marks_list"] = [old_max] * int(cfg.get("num_assignments") or 3)
                # Pad/trim lists to match num counts
                nq = int(cfg.get("num_quizzes") or 3)
                na = int(cfg.get("num_assignments") or 3)
                cfg["quiz_marks_list"]       = _pad_or_trim(cfg["quiz_marks_list"], nq, 10.0)
                cfg["assignment_marks_list"] = _pad_or_trim(cfg["assignment_marks_list"], na, 10.0)
                cfg["midterm_q_marks"]       = _pad_or_trim(
                    cfg["midterm_q_marks"], int(cfg.get("num_midterm_questions") or 2), 12.5)
                cfg["final_q_marks"]         = _pad_or_trim(
                    cfg["final_q_marks"], int(cfg.get("num_final_questions") or 3), 15.0)
                return cfg
            return _default
        except Exception as e:
            logger.exception(f"Failed to fetch AOL config: {course_uuid}")
            return _default

    @staticmethod
    def save_aol_config(course_uuid: str, cfg: dict) -> bool:
        try:
            cfg["course_id"] = course_uuid
            r = supabase.table("aol_config").upsert(cfg, on_conflict="course_id").execute()
            UProService.clear_cache()
            return bool(r.data)
        except Exception as e:
            logger.exception(f"Failed to save AOL config: {course_uuid}")
            return False

    # ── AOL Generation ────────────────────────────────────────────

    @staticmethod
    def generate_aol(
        course_uuid: str,
        components: list[str],   # e.g. ['quiz','assignment','midterm','final']
    ) -> tuple[bool, str, int]:
        """
        Generate AOL Gradebook entries from UPro scores.
        Returns (success, message, count_generated).
        """
        try:
            scheme    = GradingService.get_effective_scheme(course_uuid)
            cfg       = UProService.get_aol_config(course_uuid)
            scores    = UProService.get_upro_scores(course_uuid)

            if not scores:
                return False, "No UPro scores found. Enter scores first.", 0

            w_quiz = float(scheme["weight_quiz"])
            w_asgn = float(scheme["weight_assignment"])
            w_mid  = float(scheme["weight_midterm"])
            w_fin  = float(scheme["weight_final"])

            rows = []
            for s in scores:
                sid          = s["student_id"]
                syndicate_id = s.get("syndicate_id")

                quiz_breakdown       = []
                assignment_breakdown = []
                midterm_breakdown    = []
                final_breakdown      = []
                quiz_total = asgn_total = mid_total = fin_total = None

                # ── Quiz ──────────────────────────────────────
                # quiz_score from UPro is "obtained out of weight_quiz"
                # We distribute it proportionally across individual quizzes
                if "quiz" in components and s.get("quiz_score") is not None:
                    upro_val   = float(s["quiz_score"])          # e.g. 15 out of weight_quiz
                    q_maxs     = cfg["quiz_marks_list"]          # e.g. [10, 10, 10]
                    total_max  = sum(q_maxs)                     # e.g. 30
                    # Scale: upro_val is out of w_quiz; convert to out of total_max
                    scaled     = (upro_val / w_quiz) * total_max if w_quiz > 0 else 0.0
                    scaled     = min(scaled, total_max)
                    dist       = _distribute_marks(scaled, len(q_maxs), q_maxs)
                    quiz_breakdown = [
                        {"quiz_no": i+1, "max_marks": q_maxs[i], "obtained": dist[i]}
                        for i in range(len(q_maxs))
                    ]
                    quiz_total  = round(sum(d["obtained"] for d in quiz_breakdown), 2)
                    # Store the upro score (out of w_quiz) for display
                    quiz_upro_score = upro_val

                # ── Assignment ────────────────────────────────
                if "assignment" in components and s.get("assignment_score") is not None:
                    upro_val   = float(s["assignment_score"])
                    a_maxs     = cfg["assignment_marks_list"]
                    total_max  = sum(a_maxs)
                    scaled     = (upro_val / w_asgn) * total_max if w_asgn > 0 else 0.0
                    scaled     = min(scaled, total_max)
                    dist       = _distribute_marks(scaled, len(a_maxs), a_maxs)
                    assignment_breakdown = [
                        {"assignment_no": i+1, "max_marks": a_maxs[i], "obtained": dist[i]}
                        for i in range(len(a_maxs))
                    ]
                    asgn_total  = round(sum(d["obtained"] for d in assignment_breakdown), 2)
                    asgn_upro_score = upro_val

                # ── Midterm ───────────────────────────────────
                # midterm_score from UPro is direct total obtained (out of w_mid)
                if "midterm" in components and s.get("midterm_score") is not None:
                    upro_val   = float(s["midterm_score"])
                    q_maxs     = cfg["midterm_q_marks"]
                    total_max  = sum(q_maxs)
                    scaled     = (upro_val / w_mid) * total_max if w_mid > 0 else 0.0
                    scaled     = min(scaled, total_max)
                    dist       = _distribute_marks(scaled, len(q_maxs), q_maxs)
                    midterm_breakdown = [
                        {"question_no": i+1, "max_marks": float(q_maxs[i]), "obtained": dist[i]}
                        for i in range(len(q_maxs))
                    ]
                    mid_total   = round(sum(d["obtained"] for d in midterm_breakdown), 2)

                # ── Final ─────────────────────────────────────
                if "final" in components and s.get("final_score") is not None:
                    upro_val   = float(s["final_score"])
                    q_maxs     = cfg["final_q_marks"]
                    total_max  = sum(q_maxs)
                    scaled     = (upro_val / w_fin) * total_max if w_fin > 0 else 0.0
                    scaled     = min(scaled, total_max)
                    dist       = _distribute_marks(scaled, len(q_maxs), q_maxs)
                    final_breakdown = [
                        {"question_no": i+1, "max_marks": float(q_maxs[i]), "obtained": dist[i]}
                        for i in range(len(q_maxs))
                    ]
                    fin_total   = round(sum(d["obtained"] for d in final_breakdown), 2)

                # ── Grand total & grade ───────────────────────
                parts = [x for x in [
                    s.get("quiz_score")       if "quiz"       in components else None,
                    s.get("assignment_score") if "assignment" in components else None,
                    s.get("midterm_score")    if "midterm"    in components else None,
                    s.get("final_score")      if "final"      in components else None,
                ] if x is not None]
                grand = round(sum(float(x) for x in parts), 2) if parts else None
                letter, gpa = score_to_letter(grand, scheme) if grand is not None \
                              else (None, None)

                rows.append({
                    "course_id":             course_uuid,
                    "student_id":            sid,
                    "syndicate_id":          syndicate_id,
                    "quiz_breakdown":        json.dumps(quiz_breakdown),
                    "assignment_breakdown":  json.dumps(assignment_breakdown),
                    "midterm_breakdown":     json.dumps(midterm_breakdown),
                    "final_breakdown":       json.dumps(final_breakdown),
                    "quiz_total":            quiz_total,
                    "assignment_total":      asgn_total,
                    "midterm_total":         mid_total,
                    "final_total":           fin_total,
                    "grand_total":           grand,
                    "letter_grade":          letter,
                    "gpa_points":            gpa,
                    "status":                "draft",
                })

            if rows:
                supabase.table("aol_gradebook").upsert(rows, on_conflict="course_id,student_id").execute()
                UProService.clear_cache()

            return True, f"AOL Gradebook generated for {len(rows)} student(s).", len(rows)

        except Exception as e:
            logger.exception(f"AOL generation failed: {course_uuid}")
            return False, str(e), 0

    # ── AOL Workflow ──────────────────────────────────────────────

    @staticmethod
    def get_aol_gradebook(course_uuid: str) -> list:
        try:
            r = supabase.table("aol_gradebook")\
                .select("*, profiles(id, full_name, first_name, last_name, enrollment_number, program), syndicates(name)")\
                .eq("course_id", course_uuid)\
                .order("grand_total", desc=True)\
                .execute()
            rows = r.data or []
            # Parse JSON fields
            for row in rows:
                for field in ["quiz_breakdown","assignment_breakdown",
                               "midterm_breakdown","final_breakdown"]:
                    if isinstance(row.get(field), str):
                        try:
                            row[field] = json.loads(row[field])
                        except Exception:
                            row[field] = []
            return rows
        except Exception as e:
            logger.exception(f"Failed to fetch AOL gradebook: {course_uuid}")
            return []

    @staticmethod
    def get_student_aol_grades(student_id: str) -> list:
        """Released AOL grades for a student."""
        try:
            r = supabase.table("aol_gradebook")\
                .select("*, courses(name, code, course_id, credits)")\
                .eq("student_id", student_id)\
                .eq("status", "released")\
                .execute()
            return r.data or []
        except Exception:
            return []

    @staticmethod
    def submit_aol(course_uuid: str) -> bool:
        from datetime import datetime, timezone
        try:
            supabase.table("aol_gradebook").update({
                "status": "submitted",
                "submitted_at": datetime.now(timezone.utc).isoformat(),
            }).eq("course_id", course_uuid).eq("status", "draft").execute()
            UProService.clear_cache()
            return True
        except Exception as e:
            logger.exception(f"AOL submit failed: {course_uuid}")
            return False

    @staticmethod
    def approve_aol(course_uuid: str) -> bool:
        from datetime import datetime, timezone
        try:
            supabase.table("aol_gradebook").update({
                "status": "approved",
                "approved_at": datetime.now(timezone.utc).isoformat(),
            }).eq("course_id", course_uuid).eq("status", "submitted").execute()
            UProService.clear_cache()
            return True
        except Exception as e:
            logger.exception(f"AOL approve failed: {course_uuid}")
            return False

    @staticmethod
    def release_aol(course_uuid: str) -> bool:
        from datetime import datetime, timezone
        try:
            supabase.table("aol_gradebook").update({
                "status": "released",
                "released_at": datetime.now(timezone.utc).isoformat(),
            }).eq("course_id", course_uuid).eq("status", "approved").execute()
            UProService.clear_cache()
            return True
        except Exception as e:
            logger.exception(f"AOL release failed: {course_uuid}")
            return False

    @staticmethod
    def push_to_main_gradebook(course_uuid: str) -> tuple[bool, str]:
        """
        Copies AOL compiled grades into the regular compiled_grades table.
        Only copies rows with status='approved' or 'released'.
        """
        from datetime import datetime, timezone
        try:
            aol_rows = supabase.table("aol_gradebook").select("*")\
                .eq("course_id", course_uuid)\
                .in_("status", ["approved","released"])\
                .execute()

            if not aol_rows.data:
                return False, "No approved/released AOL grades to push."

            main_rows = []
            for row in aol_rows.data:
                main_rows.append({
                    "course_id":        course_uuid,
                    "student_id":       row["student_id"],
                    "quiz_score":       row.get("quiz_total"),
                    "assignment_score": row.get("assignment_total"),
                    "midterm_score":    row.get("midterm_total"),
                    "final_score":      row.get("final_total"),
                    "total_score":      row.get("grand_total"),
                    "letter_grade":     row.get("letter_grade"),
                    "gpa_points":       row.get("gpa_points"),
                    "status":           row.get("status","approved"),
                })

            supabase.table("compiled_grades").upsert(main_rows, on_conflict="course_id,student_id").execute()

            # Mark as pushed
            supabase.table("aol_gradebook").update({
                "pushed_to_main": True,
                "pushed_at": datetime.now(timezone.utc).isoformat(),
            }).eq("course_id", course_uuid).execute()

            UProService.clear_cache()
            return True, f"Pushed {len(main_rows)} grade(s) to main gradebook."

        except Exception as e:
            logger.exception(f"Push to main failed: {course_uuid}")
            return False, str(e)

    # ── Excel Export ──────────────────────────────────────────────

    @staticmethod
    def export_aol_to_excel(course_uuid: str, course_info: dict) -> bytes | None:
        """Generates an Excel file of the AOL gradebook."""
        try:
            import io
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            rows = UProService.get_aol_gradebook(course_uuid)
            if not rows:
                return None

            cfg = UProService.get_aol_config(course_uuid)
            wb  = openpyxl.Workbook()

            # ── Summary sheet ──────────────────────────────────
            ws = wb.active
            ws.title = "AOL Summary"

            header_fill = PatternFill("solid", fgColor="1F4E79")
            sub_fill    = PatternFill("solid", fgColor="2E75B6")
            alt_fill    = PatternFill("solid", fgColor="D6E4F0")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            bold_font   = Font(bold=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'),  bottom=Side(style='thin')
            )

            # Course info header
            ws.merge_cells("A1:J1")
            ws["A1"] = f"AOL Gradebook — {course_info.get('code','')} {course_info.get('name','')}"
            ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
            ws["A1"].alignment = Alignment(horizontal="center")

            ws.merge_cells("A2:J2")
            ws["A2"] = f"Course ID: {course_info.get('course_id','—')} | Semester: {course_info.get('semester','—')}"
            ws["A2"].alignment = Alignment(horizontal="center")
            ws["A2"].font = Font(italic=True, size=10)

            # Column headers (row 4)
            n_quiz  = cfg["num_quizzes"]
            n_asgn  = cfg["num_assignments"]
            n_mid   = cfg["num_midterm_questions"]
            n_fin   = cfg["num_final_questions"]

            headers = [
                "Enrollment No", "Student Name", "Syndicate",
                "Program",
            ]
            for i in range(n_quiz):
                headers.append(f"Quiz {i+1} (/{cfg['quiz_max_marks']})")
            headers.append("Quiz Total")
            for i in range(n_asgn):
                headers.append(f"Asgn {i+1} (/{cfg['assignment_max_marks']})")
            headers.append("Assignment Total")
            for i in range(n_mid):
                headers.append(f"Mid Q{i+1} (/{cfg['midterm_q_marks'][i] if i < len(cfg['midterm_q_marks']) else '?'})")
            headers.append("Midterm Total")
            for i in range(n_fin):
                headers.append(f"Final Q{i+1} (/{cfg['final_q_marks'][i] if i < len(cfg['final_q_marks']) else '?'})")
            headers.append("Final Total")
            headers += ["Grand Total", "Letter Grade", "GPA"]

            header_row = 4
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_idx, value=h)
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.border    = thin_border

            # Data rows
            grade_colours_hex = {
                "A": "00B050", "A-": "00B050",
                "B+": "0070C0", "B": "0070C0", "B-": "0070C0",
                "C+": "FFC000", "C": "FFC000", "C-": "FFC000",
                "D+": "FF6600", "D": "FF6600",
                "F": "FF0000",
            }

            for row_idx, g in enumerate(rows, header_row + 1):
                p      = g.get("profiles", {}) or {}
                syn    = (g.get("syndicates") or {}).get("name", "—")
                letter = g.get("letter_grade") or "—"
                fill   = alt_fill if row_idx % 2 == 0 else PatternFill()

                row_data = [
                    p.get("enrollment_number", "—"),
                    (p.get("full_name") or f"{p.get('first_name','')} {p.get('last_name','')}").strip() or "—",
                    syn,
                    p.get("program", "—"),
                ]

                # Quiz breakdown
                qb = g.get("quiz_breakdown") or []
                for i in range(n_quiz):
                    row_data.append(qb[i]["obtained"] if i < len(qb) else "—")
                row_data.append(g.get("quiz_total", "—"))

                # Assignment breakdown
                ab = g.get("assignment_breakdown") or []
                for i in range(n_asgn):
                    row_data.append(ab[i]["obtained"] if i < len(ab) else "—")
                row_data.append(g.get("assignment_total", "—"))

                # Midterm breakdown
                mb = g.get("midterm_breakdown") or []
                for i in range(n_mid):
                    row_data.append(mb[i]["obtained"] if i < len(mb) else "—")
                row_data.append(g.get("midterm_total", "—"))

                # Final breakdown
                fb = g.get("final_breakdown") or []
                for i in range(n_fin):
                    row_data.append(fb[i]["obtained"] if i < len(fb) else "—")
                row_data.append(g.get("final_total", "—"))

                row_data += [
                    round(g["grand_total"], 2) if g.get("grand_total") else "—",
                    letter,
                    g.get("gpa_points", "—"),
                ]

                for col_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border    = thin_border
                    cell.alignment = Alignment(horizontal="center")
                    if fill.fgColor and fill.fgColor.rgb != "00000000":
                        cell.fill = fill
                    # Colour letter grade
                    if col_idx == len(row_data) - 1 and letter in grade_colours_hex:
                        cell.font = Font(bold=True,
                                          color=grade_colours_hex.get(letter, "000000"))

            # Auto-fit columns
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=8)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)

            # ── Grade Distribution sheet ──────────────────────
            ws2 = wb.create_sheet("Grade Distribution")
            ws2["A1"] = "Letter Grade"
            ws2["B1"] = "Count"
            ws2["A1"].font = bold_font
            ws2["B1"].font = bold_font

            from collections import Counter
            grade_dist = Counter(g.get("letter_grade","—") for g in rows)
            for i, (grade, count) in enumerate(sorted(grade_dist.items()), 2):
                ws2.cell(row=i, column=1, value=grade)
                ws2.cell(row=i, column=2, value=count)

            out = io.BytesIO()
            wb.save(out)
            return out.getvalue()

        except Exception as e:
            logger.exception(f"Excel export failed: {course_uuid}")
            return None
